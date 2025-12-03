# ===================== SISTEMA DE MENSAGENS PROGRAMADAS (POR CLIENTES) =====================
# Autor: 2D Consultores - Integração Evolution API
# Descrição: Gerencia clientes via Excel (clientes.xlsx) e envio de mensagens programadas (SQLite)
# Requisitos: Flask, pandas, openpyxl, werkzeug, requests

import os
import uuid
import json
import sqlite3
import logging
import requests
from datetime import datetime
from typing import Tuple, List, Dict, Optional

from flask import Blueprint, request, jsonify
from werkzeug.utils import secure_filename

# Excel support
try:
    import pandas as pd
except Exception:
    pd = None

try:
    from openpyxl import Workbook, load_workbook
except Exception:
    Workbook = None
    load_workbook = None

# ===================== CONFIGURAÇÕES =====================
logger = logging.getLogger(__name__)
UPLOAD_FOLDER = 'uploads/mensagens_programadas'
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'webp'}
DATABASE = 'reunioes.db'
EXCEL_ARQUIVO = 'clientes.xlsx'

os.makedirs(UPLOAD_FOLDER, exist_ok=True)


# ===================== EVOLUTION API MANAGER =====================
class EvolutionAPIManager:
    """Gerencia a comunicação com a Evolution API (WhatsApp)."""
    def __init__(self, base_url: str, api_key: str, session: str = "default"):
        self.base_url = base_url.rstrip('/')
        self.api_key = api_key
        self.session = session
        self.headers = {"apikey": self.api_key, "Content-Type": "application/json"}

    def normalize_phone_number(self, phone: str) -> str:
        """Formata número para padrão brasileiro com DDI."""
        import re
        clean = re.sub(r'\D', '', phone or '')
        if not clean.startswith("55") and len(clean) == 11:
            clean = f"55{clean}"
        return clean

    def check_connection_status(self) -> Tuple[bool, str]:
        """Verifica se o WhatsApp está conectado."""
        try:
            # Tenta endpoint de status da instância específica primeiro
            url = f"{self.base_url}/instance/connectionState/{self.session}"
            logger.info(f"Verificando conexão em: {url}")
            
            resp = requests.get(url, headers=self.headers, timeout=8)
            
            if resp.status_code == 200:
                data = resp.json()
                logger.info(f"Resposta connectionState: {data}")
                
                # Verifica o estado na resposta
                state = data.get("state") or data.get("instance", {}).get("state", "unknown")
                is_connected = state in ["open", "connected"]
                
                if is_connected:
                    return True, state
            
            # Se falhar, tenta o endpoint fetchInstances
            url_fetch = f"{self.base_url}/instance/fetchInstances"
            logger.info(f"Tentando fetchInstances em: {url_fetch}")
            
            resp = requests.get(url_fetch, headers=self.headers, timeout=8)
            
            if resp.status_code == 200:
                data = resp.json()
                logger.info(f"Resposta fetchInstances: {json.dumps(data, indent=2)}")
                
                # Se data for uma lista
                if isinstance(data, list):
                    for inst in data:
                        # Tenta diferentes formatos de resposta da API
                        instance_name = (
                            inst.get("instanceName") or 
                            inst.get("instance", {}).get("instanceName") or
                            inst.get("name")
                        )
                        
                        logger.info(f"Comparando: '{instance_name}' com '{self.session}'")
                        
                        if instance_name == self.session:
                            # Verifica diferentes formatos de estado
                            state = (
                                inst.get("state") or 
                                inst.get("instance", {}).get("state") or
                                inst.get("status", "unknown")
                            )
                            logger.info(f"Estado encontrado: {state}")
                            is_connected = state in ["open", "connected", "Open", "Connected"]
                            return is_connected, state
                
                # Se data for um objeto
                elif isinstance(data, dict):
                    instance_name = (
                        data.get("instanceName") or 
                        data.get("instance", {}).get("instanceName") or
                        data.get("name")
                    )
                    
                    if instance_name == self.session:
                        state = (
                            data.get("state") or 
                            data.get("instance", {}).get("state") or
                            data.get("status", "unknown")
                        )
                        is_connected = state in ["open", "connected", "Open", "Connected"]
                        return is_connected, state
                
                # Instância não encontrada - lista as disponíveis
                if isinstance(data, list):
                    available = [
                        inst.get("instanceName") or inst.get("name") 
                        for inst in data
                    ]
                    logger.warning(f"Instâncias disponíveis: {available}")
                    return False, f"Instância '{self.session}' não encontrada. Disponíveis: {', '.join(available)}"
                
                return False, f"Instância '{self.session}' não encontrada"
            
            return False, f"Erro HTTP {resp.status_code}"
            
        except requests.exceptions.Timeout:
            logger.exception("Timeout ao verificar status")
            return False, "Timeout na conexão com Evolution API"
        except requests.exceptions.ConnectionError:
            logger.exception("Erro de conexão")
            return False, "Não foi possível conectar com Evolution API"
        except Exception as e:
            logger.exception("Erro inesperado ao verificar status")
            return False, str(e)

    def send_message(self, phone: str, message: str, image_path: Optional[str] = None) -> Tuple[bool, str]:
        """
        Envia texto ou imagem via Evolution API - VERSÃO CORRIGIDA COMPLETA
        Suporta envio automático de reuniões E mensagens programadas
        """
        try:
            phone_formatted = self.normalize_phone_number(phone)
            
            if image_path and os.path.exists(image_path):
                # ========== ENVIO DE IMAGEM (CORRIGIDO) ==========
                url = f"{self.base_url}/message/sendMedia/{self.session}"
                logger.info(f"📤 Enviando IMAGEM para {phone_formatted}")
                logger.info(f"   📄 Arquivo: {image_path} ({os.path.getsize(image_path)} bytes)")
                
                # Detecta MIME type correto
                ext = image_path.lower().split('.')[-1]
                mime_types = {
                    'png': 'image/png',
                    'jpg': 'image/jpeg',
                    'jpeg': 'image/jpeg',
                    'gif': 'image/gif',
                    'webp': 'image/webp'
                }
                mime_type = mime_types.get(ext, 'image/jpeg')
                
                # FORMATO CORRETO PARA IMAGEM
                with open(image_path, 'rb') as img_file:
                    files = {
                        'mediaMessage': (  # ✅ Nome correto: 'mediaMessage'
                            os.path.basename(image_path),
                            img_file,
                            mime_type
                        )
                    }
                    
                    data = {
                        'number': phone_formatted,
                        'caption': message,
                        'mediatype': 'image'  # ✅ Campo obrigatório!
                    }
                    
                    headers = {"apikey": self.api_key}  # SEM Content-Type!
                    
                    resp = requests.post(url, headers=headers, data=data, files=files, timeout=30)
                
            else:
                # ========== ENVIO DE TEXTO (FORMATO V2) ==========
                url = f"{self.base_url}/message/sendText/{self.session}"
                logger.info(f"📤 Enviando TEXTO para {phone_formatted}")
                
                # FORMATO CORRETO PARA TEXTO
                payload = {
                    "number": phone_formatted,
                    "textMessage": {  # ✅ Formato V2
                        "text": message
                    }
                }
                
                headers = {
                    "apikey": self.api_key,
                    "Content-Type": "application/json"
                }
                
                logger.info(f"   📦 Payload: {json.dumps(payload)[:100]}...")
                
                resp = requests.post(url, json=payload, headers=headers, timeout=10)

            # ========== PROCESSA RESPOSTA ==========
            logger.info(f"📊 Status: {resp.status_code}")
            
            try:
                response_json = resp.json()
                logger.info(f"📄 Resposta: {json.dumps(response_json, indent=2)[:300]}...")
            except:
                logger.info(f"📄 Resposta: {resp.text[:300]}...")

            if resp.status_code in [200, 201]:
                logger.info(f"✅ Mensagem enviada com sucesso!")
                return True, "Mensagem enviada com sucesso"
            else:
                error_msg = f"Erro HTTP {resp.status_code}: {resp.text[:200]}"
                logger.error(error_msg)
                return False, error_msg
                
        except requests.exceptions.Timeout:
            logger.exception("Timeout ao enviar")
            return False, "Timeout ao enviar mensagem"
        except requests.exceptions.ConnectionError:
            logger.exception("Erro de conexão")
            return False, "Erro de conexão"
        except FileNotFoundError:
            logger.exception("Arquivo não encontrado")
            return False, f"Arquivo não encontrado: {image_path}"
        except Exception as e:
            logger.exception("Erro ao enviar")
            return False, f"Erro: {str(e)}"

# ===================== UTILITÁRIOS EXCEL =====================
def ensure_excel() -> Tuple[bool, str]:
    """Garante que o arquivo Excel existe com a estrutura correta."""
    if pd is None or load_workbook is None or Workbook is None:
        return False, "Dependências de Excel ausentes. Instale: pip install pandas openpyxl"

    if not os.path.isfile(EXCEL_ARQUIVO):
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "CLIENTES"
            ws.append(["NOME", "EMPRESA", "WHATSAPP"])
            wb.save(EXCEL_ARQUIVO)
            logger.info(f"Arquivo {EXCEL_ARQUIVO} criado com sucesso")
        except Exception as e:
            logger.exception("Erro ao criar Excel")
            return False, f"Erro criando '{EXCEL_ARQUIVO}': {e}"
    return True, "OK"


def read_clientes() -> Tuple[bool, str, List[Dict]]:
    """Lê todos os clientes do Excel."""
    ok, msg = ensure_excel()
    if not ok:
        return False, msg, []

    try:
        df = pd.read_excel(EXCEL_ARQUIVO, dtype=str).fillna("")
        df.columns = [c.upper().strip() for c in df.columns]
        
        # Verifica se todas as colunas necessárias existem
        for col in ["NOME", "EMPRESA", "WHATSAPP"]:
            if col not in df.columns:
                return False, "Planilha precisa ter colunas: NOME, EMPRESA, WHATSAPP", []

        clientes = []
        for idx, row in df.iterrows():
            clientes.append({
                "id": int(idx),
                "nome": (row.get("NOME") or "").strip(),
                "empresa": (row.get("EMPRESA") or "").strip(),
                "whatsapp": (row.get("WHATSAPP") or "").strip(),
                "whatzap": (row.get("WHATSAPP") or "").strip()  # Alias para compatibilidade com frontend
            })
        return True, "OK", clientes
    except Exception as e:
        logger.exception("Erro lendo Excel")
        return False, str(e), []


def append_cliente_excel(nome: str, empresa: str, whatsapp: str) -> Tuple[bool, str, Optional[int]]:
    """Adiciona novo cliente ao Excel."""
    ok, msg = ensure_excel()
    if not ok:
        return False, msg, None
    try:
        wb = load_workbook(EXCEL_ARQUIVO)
        ws = wb.active
        ws.append([nome, empresa, whatsapp])
        wb.save(EXCEL_ARQUIVO)
        logger.info(f"Cliente {nome} adicionado ao Excel")
        return True, "Cliente adicionado", ws.max_row - 2
    except Exception as e:
        logger.exception("Erro ao escrever no Excel")
        return False, str(e), None


# ===================== CORE =====================
class MensagemClientes:
    def __init__(self, evolution_manager):
        self.evolution_manager = evolution_manager
        self.init_database()
        ensure_excel()

    def _conn(self):
        return sqlite3.connect(DATABASE)

    def init_database(self):
        """Inicializa as tabelas do banco de dados."""
        with self._conn() as conn:
            c = conn.cursor()
            c.execute('''
                CREATE TABLE IF NOT EXISTS mensagens_programadas (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    titulo TEXT NOT NULL,
                    texto TEXT NOT NULL,
                    imagem_path TEXT,
                    data_criacao DATETIME DEFAULT CURRENT_TIMESTAMP,
                    data_envio DATETIME,
                    status TEXT DEFAULT 'pendente',
                    total_destinatarios INTEGER DEFAULT 0,
                    total_enviados INTEGER DEFAULT 0,
                    total_erros INTEGER DEFAULT 0
                )
            ''')
            c.execute('''
                CREATE TABLE IF NOT EXISTS logs_mensagens_programadas (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    mensagem_id INTEGER NOT NULL,
                    nome_destinatario TEXT,
                    telefone TEXT,
                    status TEXT NOT NULL,
                    erro TEXT,
                    data_envio DATETIME DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (mensagem_id) REFERENCES mensagens_programadas (id)
                )
            ''')
            conn.commit()

    def salvar_imagem(self, file) -> Tuple[bool, str, Optional[str]]:
        """Salva imagem enviada e retorna o caminho."""
        if not file or not file.filename:
            return False, "Nenhum arquivo enviado", None
        
        filename = secure_filename(file.filename)
        if not filename:
            return False, "Nome de arquivo inválido", None
        
        # Verifica extensão
        ext = filename.rsplit('.', 1)[1].lower() if '.' in filename else ''
        if ext not in ALLOWED_EXTENSIONS:
            return False, f"Extensão não permitida. Use: {', '.join(ALLOWED_EXTENSIONS)}", None
        
        # Gera nome único
        unique_filename = f"{uuid.uuid4().hex}_{filename}"
        filepath = os.path.join(UPLOAD_FOLDER, unique_filename)
        
        try:
            file.save(filepath)
            logger.info(f"Imagem salva: {filepath}")
            return True, "Imagem salva", filepath
        except Exception as e:
            logger.exception("Erro ao salvar imagem")
            return False, str(e), None

    def criar_mensagem(self, titulo: str, texto: str, total_destinatarios: int, imagem_path: Optional[str] = None):
        """Cria uma nova mensagem programada."""
        with self._conn() as conn:
            c = conn.cursor()
            c.execute('''
                INSERT INTO mensagens_programadas
                    (titulo, texto, imagem_path, total_destinatarios)
                VALUES (?, ?, ?, ?)
            ''', (titulo, texto, imagem_path, total_destinatarios))
            conn.commit()
            msg_id = c.lastrowid
            logger.info(f"Mensagem {msg_id} criada: {titulo}")
            return True, "Mensagem criada", msg_id

    def enviar_mensagem(self, mensagem_id: int, destinatarios: List[Dict]):
        """Envia mensagem para lista de destinatários."""
        with self._conn() as conn:
            c = conn.cursor()
            c.execute('SELECT titulo, texto, imagem_path FROM mensagens_programadas WHERE id=?', (mensagem_id,))
            row = c.fetchone()
            if not row:
                return False, "Mensagem não encontrada", {}
            titulo, texto, imagem_path = row

            # Verifica conexão
            connected, status = self.evolution_manager.check_connection_status()
            if not connected:
                logger.warning(f"WhatsApp não conectado. Status: {status}")
                return False, f"WhatsApp não conectado ({status})", {}

            stats = {"total": len(destinatarios), "sucesso": 0, "falha": 0, "detalhes": []}
            
            for d in destinatarios:
                nome = d.get("nome", "")
                # Aceita tanto 'whatsapp' quanto 'whatzap'
                telefone = self.evolution_manager.normalize_phone_number(
                    d.get("whatsapp") or d.get("whatzap", "")
                )
                
                # Substitui {nome} na mensagem
                msg = f"*{titulo}*\n\n{texto.replace('{nome}', nome)}"
                
                # Envia mensagem
                success, result = self.evolution_manager.send_message(telefone, msg, imagem_path)
                
                if success:
                    stats["sucesso"] += 1
                    status_envio, erro = "success", None
                    logger.info(f"Mensagem enviada para {nome} ({telefone})")
                else:
                    stats["falha"] += 1
                    status_envio, erro = "error", result
                    stats["detalhes"].append(f"{nome}: {result}")
                    logger.error(f"Falha ao enviar para {nome} ({telefone}): {result}")

                # Registra log
                c.execute('''
                    INSERT INTO logs_mensagens_programadas
                        (mensagem_id, nome_destinatario, telefone, status, erro)
                    VALUES (?, ?, ?, ?, ?)
                ''', (mensagem_id, nome, telefone, status_envio, erro))
            
            conn.commit()

            # Define status final
            if stats["falha"] == 0:
                status_final = "enviada"
            elif stats["sucesso"] > 0:
                status_final = "enviada_com_erros"
            else:
                status_final = "falhou"

            # Atualiza mensagem
            c.execute('''
                UPDATE mensagens_programadas
                   SET status=?, total_enviados=?, total_erros=?, data_envio=?
                 WHERE id=?
            ''', (status_final, stats["sucesso"], stats["falha"], datetime.now().isoformat(), mensagem_id))
            conn.commit()

            logger.info(f"Envio concluído. Sucesso: {stats['sucesso']}/{stats['total']}")
            return True, f"Enviadas {stats['sucesso']}/{stats['total']}", stats

    def listar_mensagens(self):
        """Lista todas as mensagens programadas com formato compatível com frontend."""
        with self._conn() as conn:
            c = conn.cursor()
            c.execute('SELECT * FROM mensagens_programadas ORDER BY data_criacao DESC LIMIT 50')
            cols = [d[0] for d in c.description]
            mensagens = []
            for row in c.fetchall():
                msg_dict = dict(zip(cols, row))
                # Adiciona campos esperados pelo frontend
                msg_dict['tem_imagem'] = bool(msg_dict.get('imagem_path'))
                msg_dict['created_at'] = msg_dict.get('data_criacao')
                msg_dict['sent_at'] = msg_dict.get('data_envio')
                mensagens.append(msg_dict)
            return mensagens

    def obter_logs_mensagem(self, mensagem_id: int):
        """Obtém logs de envio de uma mensagem específica."""
        with self._conn() as conn:
            c = conn.cursor()
            c.execute('''
                SELECT nome_destinatario, telefone, status, erro, data_envio 
                FROM logs_mensagens_programadas 
                WHERE mensagem_id=?
                ORDER BY data_envio DESC
            ''', (mensagem_id,))
            cols = [d[0] for d in c.description]
            logs = []
            for row in c.fetchall():
                log_dict = dict(zip(cols, row))
                # Adiciona campos esperados pelo frontend
                log_dict['nome'] = log_dict.get('nome_destinatario')
                log_dict['sent_at'] = log_dict.get('data_envio')
                logs.append(log_dict)
            return logs

    def dashboard_stats(self):
        """Retorna estatísticas do dashboard."""
        with self._conn() as conn:
            c = conn.cursor()
            c.execute('SELECT COUNT(*) FROM mensagens_programadas')
            total = c.fetchone()[0]
            c.execute("SELECT COUNT(*) FROM mensagens_programadas WHERE status IN ('enviada', 'enviada_com_erros')")
            enviadas = c.fetchone()[0]
        ok, _, clientes = read_clientes()
        return {
            "total_mensagens": total,
            "total_enviadas": enviadas,
            "total_clientes": len(clientes) if ok else 0
        }


# ===================== BLUEPRINT/ROTAS =====================
def create_mensagens_clientes_blueprint(evolution_manager):
    """Cria e configura o blueprint Flask."""
    svc = MensagemClientes(evolution_manager)
    bp = Blueprint('mensagens_clientes', __name__, url_prefix='/api/clientes-msg')

    @bp.get('/dashboard')
    def dashboard():
        """Retorna estatísticas do dashboard."""
        try:
            return jsonify(success=True, stats=svc.dashboard_stats())
        except Exception as e:
            logger.exception("Erro no dashboard")
            return jsonify(success=False, message=str(e)), 500

    @bp.get('/clientes')
    def clientes_list():
        """Lista todos os clientes do Excel."""
        try:
            ok, msg, clientes = read_clientes()
            return jsonify(success=ok, message=msg, clientes=clientes)
        except Exception as e:
            logger.exception("Erro ao listar clientes")
            return jsonify(success=False, message=str(e), clientes=[]), 500

    @bp.post('/clientes/adicionar')
    def clientes_add():
        """Adiciona novo cliente ao Excel."""
        try:
            data = request.get_json(force=True)
            nome = data.get("nome", "").strip()
            empresa = data.get("empresa", "").strip()
            # Aceita tanto 'whatsapp' quanto 'whatzap'
            whatsapp = data.get("whatsapp") or data.get("whatzap", "")
            whatsapp = whatsapp.strip()
            
            if not nome or not whatsapp:
                return jsonify(success=False, message="Nome e WhatsApp são obrigatórios"), 400
            
            ok, msg, _ = append_cliente_excel(nome, empresa, whatsapp)
            return jsonify(success=ok, message=msg)
        except Exception as e:
            logger.exception("Erro ao adicionar cliente")
            return jsonify(success=False, message=str(e)), 500

    @bp.post('/mensagens/enviar')
    def enviar():
        """Envia mensagem programada para destinatários selecionados."""
        try:
            titulo = request.form.get('titulo', '').strip()
            texto = request.form.get('texto', '').strip()
            destinatarios_json = request.form.get('destinatarios', '[]')
            
            try:
                selecionados = json.loads(destinatarios_json)
            except:
                return jsonify(success=False, message="Formato inválido de destinatários"), 400
            
            imagem = request.files.get('imagem')

            # Validações
            if not titulo or not texto:
                return jsonify(success=False, message="Título e texto são obrigatórios"), 400
            
            if not selecionados or len(selecionados) == 0:
                return jsonify(success=False, message="Selecione ao menos um destinatário"), 400

            imagem_path = None
            if imagem and imagem.filename:
                ok, msg, caminho = svc.salvar_imagem(imagem)
                if ok:
                    imagem_path = caminho
                else:
                    return jsonify(success=False, message=msg), 400

            ok, msg, mensagem_id = svc.criar_mensagem(titulo, texto, len(selecionados), imagem_path)
            if not ok:
                return jsonify(success=False, message=msg), 400

            ok2, msg2, stats = svc.enviar_mensagem(mensagem_id, selecionados)
            return jsonify(success=ok2, message=msg2, stats=stats)
        except Exception as e:
            logger.exception("Erro ao enviar mensagem")
            return jsonify(success=False, message=str(e)), 500

    @bp.get('/historico')
    def historico():
        """Retorna histórico de mensagens enviadas."""
        try:
            return jsonify(success=True, mensagens=svc.listar_mensagens())
        except Exception as e:
            logger.exception("Erro ao listar histórico")
            return jsonify(success=False, message=str(e), mensagens=[]), 500

    @bp.get('/mensagem/<int:mensagem_id>/logs')
    def logs(mensagem_id):
        """Retorna logs de envio de uma mensagem específica."""
        try:
            return jsonify(success=True, logs=svc.obter_logs_mensagem(mensagem_id))
        except Exception as e:
            logger.exception("Erro ao listar logs")
            return jsonify(success=False, message=str(e), logs=[]), 500

    return bp